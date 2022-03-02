import pandas as pd
from matplotlib import pyplot as plt
from sklearn.feature_extraction.text import TfidfVectorizer
import re
import jieba.posseg as psg
from openpyxl import load_workbook
from sklearn.cluster import KMeans
import numpy as np

from scipy. spatial.distance import cdist


def getcomment(i):
    wb = load_workbook("news_data.xlsx")
    sheet = wb["Sheet1"]
    comment_str =  sheet.cell(i, 1).value
    return comment_str

def fenci(string):
    counts = {}
    tingyong_set = []
    f = open('hit_stopwords.txt', encoding='utf-8')
    for line in f.readlines()[1:]:
        data = line.strip().split('\t')
        tingyong_set = tingyong_set + data
    text_lines = string.split('。')
    for line in text_lines:
        line_seg = psg.cut(line)
        for word_flag in line_seg:
            word = re.sub("[^\u4e00-\u9fa5]", "", word_flag.word)           #[^…]表示不在[]中的字符 不在中文编码的字符转换为 ‘’
            if  len(word) > 1 and word not in tingyong_set:
                counts[word] = counts.get(word, 0) + 1

    return counts

def tfidf(df):
    tfidf_model = TfidfVectorizer(max_features=100)
    tfidf_df = pd.DataFrame(tfidf_model.fit_transform(df).todense())
    tfidf_df.columns = sorted(tfidf_model.vocabulary_)
    print(tfidf_df.head())        #默认n = 5 只输出前5行
    return tfidf_df

def cluseter(tfidf_df):
    kmeans = KMeans(n_clusters=2)
    # 设置集群数量
    kmeans.fit(tfidf_df)
    # 使用fit方法训练tf-idf向量化后的文本数据
    y_kmeans = kmeans.predict(tfidf_df)
    # print(y_kmeans)

def youhua(tfidf_df):
    # 肘部法求最佳K值
    distortions = []
    K = range(1,10)
    for k in K:
        kmeanModel = KMeans(n_clusters=k)
        kmeanModel.fit(tfidf_df)

        distortions.append(sum(np.min(cdist(tfidf_df,kmeanModel.cluster_centers_, 'euclidean'),axis = 1)) / tfidf_df.shape[0])
    # 绘制肘部图形
    plt.plot(K, distortions,'bx-')
    plt.xlabel('k')
    plt.ylabel(' Distortion')
    plt.title(' The Elbow Method showing the optimal number of clusters')
    plt.show()


if __name__ == '__main__':
    count = []

    for i in range(900, 1000):
        comment_list = getcomment(i)
        counts = fenci(comment_list)
        count.append(counts)
    # print(count)
    ori = pd.read_excel('kang_word_freq.xlsx')
    feats = []
    for word in ori['word']:
        feats.append(word)

    word_freq = pd.DataFrame( count, columns= feats,dtype=float)

    word_freq = word_freq.fillna(0)
    # print(word_freq)
    tfidf_df = tfidf(word_freq)
    # cluseter(tfidf_df)
    # youhua(tfidf_df)
