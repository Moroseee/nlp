# coding:utf-8

import jieba
import jieba.posseg as pseg
import os
import sys

import pandas as pd
from matplotlib import pyplot as plt
from sklearn import feature_extraction, metrics
from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import TfidfTransformer, TfidfVectorizer
from sklearn.feature_extraction.text import CountVectorizer
from openpyxl import load_workbook
import jieba.posseg as psg
import re



if __name__ == "__main__":
    wb = load_workbook("news_data.xlsx")
    sheet = wb["Sheet1"]
    list = []
    tingyong_set = []
    f = open('hit_stopwords.txt', encoding='utf-8')
    for line in f.readlines()[1:]:
        data = line.strip().split('\t')  # ltrip 去除头   rtrip 去除尾   strip()（去空格）两头都去除   strip('0') 去除首尾字符 0
        # 得到一个列表
        tingyong_set = tingyong_set + data
    for i in range(100,2500):
        comment_str = sheet.cell(i, 1).value
        string =''
        text_lines = comment_str.split('。')
        for line in text_lines:
            line_seg = psg.cut(line)
            for word_flag in line_seg:
                word = re.sub("[^\u4e00-\u9fa5]", "", word_flag.word)
                if len(word) > 1 and word not in tingyong_set:
                    string =string +  ' '+word

        list.append(string)
    vectorizer=CountVectorizer()#该类会将文本中的词语转换为词频矩阵，矩阵元素a[i][j] 表示j词在i类文本下的词频
    transformer=TfidfTransformer()#该类会统计每个词语的tf-idf权值
    tfidf=transformer.fit_transform(vectorizer.fit_transform(list))#第一个fit_transform是计算tf-idf，第二个fit_transform是将文本转为词频矩阵
    word=vectorizer.get_feature_names()#获取词袋模型中的所有词语
    weight=tfidf.toarray()#将tf-idf矩阵抽取出来，元素a[i][j]表示j词在i类文本中的tf-idf权重
    # for i in range(len(weight)):#打印每类文本的tf-idf词语权重，第一个for遍历所有文本，第二个for便利某一类文本下的词语权重
    #     print("-------这里输出第",i,u"类文本的词语tf-idf权重------")
    #     for j in range(len(word)):
    #         print(word[j],weight[i][j])
    tfidf_model = TfidfVectorizer(max_features=100)
    tfidf_df = pd.DataFrame(tfidf_model.fit_transform(list).todense())
    tfidf_df.columns = sorted(tfidf_model.vocabulary_)
    # print(tfidf_df.head())
    kmeans = KMeans(n_clusters=3)
    # 设置集群数量
    kmeans.fit(tfidf_df)
    # 使用fit方法训练tf-idf向量化后的文本数据
    y_predict = kmeans.predict(tfidf_df)
    tfidf_df = tfidf_df.values


    #
    # plt.scatter(tfidf_df[:, 0], tfidf_df[:, 1], c=y_predict)
    # plt.show()
    # print(kmeans.predict((tfidf_df[:10, :])))
    # print(metrics.calinski_harabasz_score(tfidf_df, y_predict))
    # print(kmeans.cluster_centers_)
    # print(kmeans.inertia_)
    # print(metrics.silhouette_score(tfidf_df, y_predict))\


weight = tfidf.toarray()
from sklearn.manifold import TSNE
tsne = TSNE(n_components=2)  #维度
decomposition_data = tsne.fit_transform(weight)
x = []
y = []
for i in decomposition_data:
    x.append(i[0])
    y.append(i[1])
# 聚类结果散点图表示
fig = plt.figure(figsize=(10, 10))
ax = plt.axes()
plt.scatter(x, y, c=y_predict)#plt画图plt.scatter画散点图
plt.xticks(()) #xticks到底有什么用，其实就是想把坐标轴变成自己想要的样子
plt.yticks(())
# for i in range(len(x)):
#     plt.annotate(list[i], xy=(x[i], y[i]), xytext=(x[i] + 0.1, y[i] + 0.1))  #plt.annotate为散点图添加中文标注，这里xy是需要标记的坐标，xytext是对应的标签坐标
plt.show()
