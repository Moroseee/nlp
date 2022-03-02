from openpyxl import load_workbook
import jieba.posseg as psg
import re
import pandas as pd


def getcomment():
    #打开一个xlsx文件
    wb = load_workbook("news_data.xlsx")
    sheet = wb["Sheet1"]
    kang_comment_str = ''

    max_r = sheet.max_row
    for i in range(2,max_r):
        if sheet.cell(i,2).value == "康明斯":
        # if sheet.cell(i, 2).value == "山东重工":

            kang_comment_str = kang_comment_str + sheet.cell(i,1).value
            # comment_list=list(set(kang_comment_str))  #无重复评论

    return kang_comment_str


def fenci(string):
    counts = {}
    tingyong_set = []
    f = open('hit_stopwords.txt', encoding='utf-8')
    for line in f.readlines()[1:]:
        data = line.strip().split('\t')  # ltrip 去除头   rtrip 去除尾   strip()（去空格）两头都去除   strip('0') 去除首尾字符 0
        # 得到一个列表
        tingyong_set = tingyong_set + data
    text_lines = string.split('。')
    for line in text_lines:
        line_seg = psg.cut(line)
        for word_flag in line_seg:
            word = re.sub("[^\u4e00-\u9fa5]", "", word_flag.word)
            if  len(word) > 1 and word not in tingyong_set:
                counts[word] = counts.get(word, 0) + 1
    word_freq = pd.DataFrame({'word': list(counts.keys()), 'freq': list(counts.values())})
    word_freq = word_freq.sort_values(by='freq', ascending=False)

    word_freq = pd.DataFrame(word_freq.values.T, index=word_freq.columns, columns=word_freq.index)         #转置
    word_freq.to_excel("shan_zhuan_word_freq.xlsx", index=False)



if __name__ == '__main__':
    comment_list = getcomment()
    cut = fenci(comment_list)
    print('________')


