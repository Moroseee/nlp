import random

from openpyxl import load_workbook
import jieba.posseg as psg
import re
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import roc_curve, auc, f1_score
import matplotlib.pyplot as plt
from sklearn.linear_model import LogisticRegression
from sklearn.neural_network import MLPClassifier
import joblib   #joblib模块

def getcomment(i):

    kang_comment_str = ''
    a = 1
    if sheet.cell(i,2).value != "康明斯":
    # if sheet.cell(i, 2).value == "山东重工":
        a = 0
        # kang_comment_str = kang_comment_str + sheet.cell(i,1).value
        # comment_list=list(set(kang_comment_str))  #无重复评论
    kang_comment_str = sheet.cell(i, 1).value
    return kang_comment_str,a


def fenci(string,a):
    counts = {}
    tingyong_set = []
    list = []
    f = open('hit_stopwords.txt', encoding='utf-8')
    for line in f.readlines()[1:]:
        data = line.strip().split('\t')  # ltrip 去除头   rtrip 去除尾   strip()（去空格）两头都去除   strip('0') 去除首尾字符 0
        # 得到一个列表
        tingyong_set = tingyong_set + data
    text_lines = string.split('。')
    for line in text_lines:
        line_seg = psg.cut(line)
        for word_flag in line_seg:
            word = re.sub("[^\u4e00-\u9fa5]", "", word_flag.word)
            if  len(word) > 1 and word not in tingyong_set:
                counts[word] = counts.get(word, 0) + 1
    counts['kang'] = a
    # list.append(counts)
    # word_freq = pd.DataFrame.from_dict(counts,orient='columns')
    # word_freq = pd.DataFrame(counts,index=[0])
    # word_freq = pd.DataFrame({'word': counts.keys(), 'freq': counts.values()},dtype = 'float')
    # word_freq = word_freq.sort_values(by='freq', ascending=False)
    # word_freq = pd.DataFrame(word_freq.values.T, index=word_freq.columns, columns=word_freq.index)         #转置

    # word_freq.to_excel("shan_zhuan_word_freq.xlsx", index=False)

    return counts

def trainNB(word_freq):

    X_train, X_test, y_train, y_test = train_test_split(word_freq, word_freq[['kang']], test_size=0.2, random_state=0)
    model = GaussianNB()
    model.fit(X_train, y_train)
    joblib.dump(model, 'GaussianNB.pkl')
    print("GaussianNB",model.score(X_test,y_test))
    model = LogisticRegression()
    model.fit(X_train, y_train)
    joblib.dump(model, 'LogisticRegression.pkl')
    print("LogisticRegression",model.score(X_test, y_test))
    model = MLPClassifier(solver='lbfgs', alpha=1e-5, hidden_layer_sizes = (100, 25), random_state = 1)
    model.fit(X_train, y_train)
    joblib.dump(model, 'MLPClassifier.pkl')
    print("MLPClassifiermodel",model.score(X_test, y_test))
    from sklearn import svm
    model = svm.SVC(probability=True)
    # y_pred = gnb.fit(X_train, y_train).predict(X_test)
    model.fit(X_train, y_train)
    joblib.dump(model, 'SVC.pkl')
    print("SVC",model.score(X_test,y_test))
    # Y_pred = model.predict_proba(X_test)[:,1]
    #
    # Y_pred_lab = model.predict(X_test)
    # fpr, tpr, thresholds = roc_curve(y_test, Y_pred)
    # roc_auc = auc(fpr, tpr)
    # plt.plot([0, 1], [0, 1], 'k--')
    # plt.plot(fpr, tpr)
    # plt.title(f'auc:{roc_auc :.3f},f1 socre:{f1_score(y_test, Y_pred_lab):.3f}')
    # plt.xlabel('False positive rate')
    # plt.ylabel('True positive rate')
    # plt.show()



if __name__ == '__main__':
    wb = load_workbook("news_data.xlsx")
    sheet = wb["Sheet1"]
    max_r = sheet.max_row
    df_empty = pd.DataFrame()
    count =[]
    for i in range(2,2000):
        comment_list,a = getcomment(i)
        counts = fenci(comment_list,a)
        count.append(counts)
        # df_empty = pd.concat([df_empty, word_freq])
    print(count)
    word_freq = pd.DataFrame(count)
    # df_empty.to_excel("shan_zhuan_word_freq.xlsx", index=False)
    # print(df_empty)

    word_freq = word_freq.fillna(0)   # 将NaN值转化为0     df.replace(nan,0)也可以
    # word_freq.to_excel("test.xlsx", index=False)

    print(word_freq)
    trainNB(word_freq)




    print('________')




